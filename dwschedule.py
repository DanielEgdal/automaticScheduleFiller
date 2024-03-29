import pandas as pd
import pytz
from datetime import time
from collections import defaultdict
from openpyxl import load_workbook
import os

def getDaySplit(schedule):
	daySplit= [0] # Starting at 0, as this indicates the first event of the comp
	for i in range(1,len(schedule)):
		if schedule[i][1].day == schedule[i-1][1].day:
			pass
		else:
			daySplit.append(i)
	return daySplit

def fillInCells(sch,i,val,event_hit,eventPro,eventcount,eventCut):
	event_hit[val[0]] += 1
	sch.cell(row=i,column=3).value = val[0] # set event name
	if val[0] in eventcount: # if it is an event, where we need to add competitors
		if event_hit[val[0]] == 1: # first round
			sch.cell(row=i,column=10).value = eventcount[val[0]] # set competitors
		else: # advancement conditions for the rest
			if val[0] not in {'3x3 MBLD','3x3 FMC'}:
				if eventPro[val[0]][event_hit[val[0]]][0]: # Ranking based
					sch.cell(row=i,column=10).value = eventPro[val[0]][event_hit[val[0]]][1]
				else: # Percentage based
					sch.cell(row=i,column=10).value = int((eventPro[val[0]][event_hit[val[0]]][1]/100)*eventcount[val[0]])
		if val[0] not in {'3x3 MBLD','3x3 FMC'}:
			if eventCut[val[0]][event_hit[val[0]]] == 1: # Cumulative time limit
				sch.cell(row=i,column=8).value = eventCut[val[0]][event_hit[val[0]]][1]
			elif eventCut[val[0]][event_hit[val[0]]] == 2: # Cutoff
				sch.cell(row=i,column=9).value = eventCut[val[0]][event_hit[val[0]]][1]
	i+=1
	return i

def wallinSchedule(data,stations):
	timezone = pytz.timezone(data["schedule"]["venues"][0]["timezone"])
	# Convert to DWs names
	mapping = {'222':'2x2','333':'3x3','444':'4x4','555':'5x5','666':'6x6','777':'7x7','sq1':'Square-1'
	,'skewb':'Skewb','minx':'Megaminx', 'clock':'Clock','pyram':'Pyraminx','333oh':'3x3 OH','333fm':'3x3 FMC',
	'333bf':'3x3 BLD','444bf':'4x4 BLD','555bf':'555 BLD','333mbf':'3x3 MBLD',
	'tutorial':'Deltagarintroduktion','lunch':'Lunch','misc':'Städning','awards':'Prisutdelning','registration':'Registrering', 'checkin':'Registrering'}

	eventcount = defaultdict(int) # The amount of competitors per event

	for person in data['persons']:
		if person['registration']:
			if person['registration']['status'] == 'accepted':
				for event in person['registration']['eventIds']:
					eventcount[mapping[event]]+=1

	eventCut = defaultdict(dict) # Cutoff / Cumulative
	eventPro = defaultdict(dict) # Advancement conditions

	for event in data["events"]:
		for round in event['rounds']:
			eventname = round['id'].split('-')
			if eventname[0] not in {'333mbf','333fm'}:
				if round['timeLimit']['cumulativeRoundIds']:
					eventCut[mapping[eventname[0]]][int(eventname[1][1:])] = (1,round['timeLimit']['centiseconds']*6000)
				elif round['cutoff']:
					eventCut[mapping[eventname[0]]][int(eventname[1][1:])] = (2,round['cutoff']*6000)
				else:
					eventCut[mapping[eventname[0]]][int(eventname[1][1:])] = (0,0)

				if round['advancementCondition']:
					if round['advancementCondition']['type'] == 'ranking':
						eventPro[mapping[eventname[0]]][int(eventname[1][1:])+1] = (1,round['advancementCondition']['level'])
					elif round['advancementCondition']['type'] == 'percent':
						eventPro[mapping[eventname[0]]][int(eventname[1][1:])+1]  = (0,round['advancementCondition']['level'])
					else:
						eventPro[mapping[eventname[0]]][int(eventname[1][1:])+1]  = (0,75)

	schedule = [] 
	for room in data["schedule"]["venues"][0]['rooms']:
		for val in room["activities"]:
			starttime = pd.Timestamp(val['startTime'][:-1]).tz_localize(pytz.utc).tz_convert(timezone)
			endtime = pd.Timestamp(val['endTime'][:-1]).tz_localize(pytz.utc).tz_convert(timezone)
			ev = val['activityCode'].split('-')
			if ev[0][0] != 'o':
				ev = mapping[ev[0]] # Map event name
			else: # Activities defined by user. (non events)
				try: 
					ev=mapping[ev[1]] 
				except KeyError: # To avoid throwing an error, skip the activity
					continue
			schedule.append((ev,starttime,endtime))
	schedule.sort(key=lambda x:x[1]) # Sort based on start time

	daySplit = getDaySplit(schedule)
	print(daySplit)

	script_dir = os.path.dirname(os.path.abspath(__file__))
	filename = os.path.join(script_dir, 'dwscheduleOct2023.xlsx')

	wb = load_workbook(filename = filename) # open the schedule template

	sheet_names = wb.sheetnames
	sch = wb[sheet_names[0]] # The interesting tab
	sch.cell(row=2,column=17,value=stations)

	event_hit = defaultdict(int)

	differentDayInsertion(daySplit,schedule,sch,event_hit,eventPro,eventcount,eventCut)

	# wb.save(f"dwschedule{id}.xlsx")
	return wb

def differentDayInsertion(daySplit,schedule,sheet,event_hit,eventPro,eventcount,eventCut):
	startingPointDay = [8,31,54,80] # The first row for the day
	for idx,day in enumerate(daySplit): 
		i = startingPointDay[idx]
		continue_to = len(schedule)
		if idx+1 < len(daySplit):
			continue_to = daySplit[idx+1]
		sheet.cell(row=i-2,column=1,value=schedule[day][1].hour) # Starting time of the day
		sheet.cell(row=i-2,column=2).value = schedule[day][1].minute
		for val in schedule[daySplit[idx]:continue_to]:
			i = fillInCells(sheet,i,val,event_hit,eventPro,eventcount,eventCut)



